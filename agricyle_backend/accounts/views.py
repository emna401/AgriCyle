from __future__ import annotations
from django.core.exceptions import ValidationError
from django.contrib.auth.password_validation import validate_password
from collections import defaultdict
from datetime import datetime
from django.db.models import Count, Sum, F, DecimalField, ExpressionWrapper, Q
from django.contrib.auth import authenticate, get_user_model
from django.utils import timezone
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.core.mail import send_mail
from rest_framework.response import Response
from django.core.validators import validate_email
from rest_framework import status
from rest_framework.decorators import api_view, action, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from django.db import transaction
from rest_framework import viewsets
from typing import TYPE_CHECKING
from django.conf import settings
import io
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.template.loader import render_to_string
import secrets
import string
from datetime import datetime, timedelta
# Importations locales
from .auth_utils import generate_otp_code, hash_code, otp_expiry
from .serializers import UserSerializer
from .models import Declaration, Product, Order, OrderItem, GoodPractice, OTPCode, AdminAuditLog
from .serializers import (
    UserMeSerializer, DeclarationSerializer, ProductSerializer,
    OrderSerializer, GoodPracticeSerializer,
    LoginSerializer, ResetRequestSerializer, ResetConfirmSerializer,
    CartAddItemSerializer, CartUpdateItemSerializer, OrderSetStatusSerializer
)
from .permissions import IsNotSuspended, IsOwnerOrAdmin, IsAdminRole

if TYPE_CHECKING:
    from django.contrib.auth.models import AbstractBaseUser

User = get_user_model()

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils import timezone
import traceback
import json
RESET_TOKENS = {}

def generate_token(length=32):
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))
# ---------- Helpers ----------
def _xlsx_response(filename: str, wb: Workbook) -> HttpResponse:
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


def _admin_only(user) -> bool:
    return user.is_authenticated and user.is_admin_role()



def _apply_excel_styles(ws):
    """Applique des styles Excel de base"""
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Appliquer aux cellules
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            if cell.row == 1:  # Titre principal
                continue
            elif cell.row == 3:  # En-têtes de tableau
                if cell.value and str(cell.value).strip():
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
            else:
                cell.font = data_font
                cell.alignment = data_alignment
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
        
from rest_framework_simplejwt.tokens import RefreshToken

# ---------- Registration ----------


@api_view(['PUT'])
@permission_classes([IsAdminUser])
def admin_update_user(request, user_id):
    """Admin met à jour un utilisateur"""
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({'detail': 'Utilisateur non trouvé'}, status=status.HTTP_404_NOT_FOUND)

    # Champs autorisés pour modification par admin
    allowed_fields = [
        'username', 'email', 'phone', 'role', 'status', 'city',
        'governorate', 'delegation', 'farm_name', 'farm_address',
        'farm_lat', 'farm_lng'
    ]

    for field in allowed_fields:
        if field in request.data:
            setattr(user, field, request.data[field])

    user.save()

    return Response({
        'message': 'Utilisateur mis à jour',
        'user': {
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'role': user.role,
            'status': user.status,
            'city': user.city,
        }
    }, status=status.HTTP_200_OK)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsAdminUser])
def get_users(request):
    """Récupérer tous les utilisateurs avec filtres"""
    role_filter = request.GET.get('role')
    status_filter = request.GET.get('status')
    search = request.GET.get('search', '')
    governorate = request.GET.get('governorate')
    city_filter = request.GET.get('city')
    queryset = User.objects.all()
    
    # Filtres
    if role_filter:
        queryset = queryset.filter(role=role_filter)
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    if city_filter:  # <-- AJOUTEZ CE FILTRE POUR LA VILLE
        queryset = queryset.filter(city__icontains=city_filter)
    if search:
        queryset = queryset.filter(
            Q(username__icontains=search) |
            Q(email__icontains=search) |
            Q(phone__icontains=search) |
            Q(farm_name__icontains=search) |
            Q(city__icontains=search)
        )
    
    # Sérialisation simple
    users_data = []
    for user in queryset.order_by('-date_joined'):
        users_data.append({
            'id': user.id,
            'username': user.username,
            'email': user.email,
            'phone': user.phone,
            'role': user.role,
            'status': user.status,
            'city': user.city,
            'governorate': user.governorate,
            'delegation': user.delegation,
            'farm_name': user.farm_name,
            'farm_address': user.farm_address,
            'farm_lat': user.farm_lat,
            'farm_lng': user.farm_lng,
            'date_joined': user.date_joined,
            'last_login': user.last_login,
        })
    
    return Response(users_data)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def suspend_user(request, user_id):
    """Suspendre un utilisateur"""
    try:
        user = User.objects.get(id=user_id)
        user.status = User.Status.SUSPENDED
        user.save()
        return Response({'message': 'Utilisateur suspendu'})
    except User.DoesNotExist:
        return Response({'detail': 'Utilisateur non trouvé'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def activate_user(request, user_id):
    """Activer un utilisateur"""
    try:
        user = User.objects.get(id=user_id)
        user.status = User.Status.ACTIVE
        user.save()
        return Response({'message': 'Utilisateur activé'})
    except User.DoesNotExist:
        return Response({'detail': 'Utilisateur non trouvé'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def reset_user_password(request, user_id):
    """Admin réinitialise le mot de passe d'un utilisateur"""
    try:
        user = User.objects.get(id=user_id)
        new_password = request.data.get('new_password')
        
        if not new_password:
            return Response({'detail': 'Nouveau mot de passe requis'}, status=status.HTTP_400_BAD_REQUEST)
        
        user.set_password(new_password)
        user.save()
        
        return Response({'message': 'Mot de passe réinitialisé'})
    except User.DoesNotExist:
        return Response({'detail': 'Utilisateur non trouvé'}, status=status.HTTP_404_NOT_FOUND)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def delete_user(request, user_id):
    """Supprimer un utilisateur"""
    try:
        user = User.objects.get(id=user_id)
        
        # Empêcher la suppression d'admin
        if user.role in [User.Role.ADMIN, User.Role.SUPERADMIN]:
            return Response(
                {'detail': 'Impossible de supprimer un administrateur'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        user.delete()
        return Response({'message': 'Utilisateur supprimé'})
    except User.DoesNotExist:
        return Response({'detail': 'Utilisateur non trouvé'}, status=status.HTTP_404_NOT_FOUND)




@api_view(['GET'])
@permission_classes([IsAdminUser])
def export_users_xlsx(request):
    """Exporter les utilisateurs en Excel avec style professionnel"""
    try:
        # Récupérer les paramètres de filtrage (optionnels)
        role = request.query_params.get('role')
        status = request.query_params.get('status')
        governorate = request.query_params.get('governorate')
        delegation = request.query_params.get('delegation')
        
        # Appliquer les filtres
        users_qs = User.objects.all().order_by('-date_joined')
        
        if role:
            users_qs = users_qs.filter(role=role)
        if status:
            users_qs = users_qs.filter(status=status)
        if governorate:
            users_qs = users_qs.filter(governorate=governorate)
        if delegation:
            users_qs = users_qs.filter(delegation__icontains=delegation)
        
        # Création du workbook
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        if 'Sheet' in wb.sheetnames:
            default_ws = wb['Sheet']
            wb.remove(default_ws)
        
        # === FEUILLE 1 : LISTE DES UTILISATEURS ===
        ws_main = wb.create_sheet("Utilisateurs")
        
        # Titre principal
        ws_main.merge_cells('A1:N1')
        title_cell = ws_main['A1']
        title_cell.value = "LISTE DES UTILISATEURS - AGRI CYCLE"
        title_cell.font = Font(name='Arial', size=16, bold=True, color='2E7D32')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Informations de génération
        ws_main['A2'] = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {users_qs.count()} utilisateurs"
        ws_main.merge_cells('A2:N2')
        ws_main['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
        ws_main['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # En-têtes du tableau
        headers = [
            "ID", "NOM D'UTILISATEUR", "EMAIL", "TÉLÉPHONE", "ROLE", 
            "STATUT", "GOUVERNORAT", "DÉLÉGATION", "NOM FERME", "ADRESSE",
            "LATITUDE", "LONGITUDE", "DATE INSCRIPTION", "DERNIÈRE CONNEXION"
        ]
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_main.cell(row=3, column=col_idx, value=header)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        
        # Labels pour les rôles et statuts
        role_labels = {
            'FARMER': 'Agriculteur',
            'BENEFICIARY': 'Bénéficiaire',
            'DUAL': 'Dual',
            'ADMIN': 'Admin',
            'SUPERADMIN': 'Super Admin'
        }
        
        status_labels = {
            'ACTIVE': 'Actif',
            'SUSPENDED': 'Suspendu',
            'PENDING': 'En attente',
            'BLOCKED': 'Bloqué'
        }
        
        # Remplir les données
        row = 4
        for user in users_qs:
            ws_main.cell(row=row, column=1, value=user.id)
            ws_main.cell(row=row, column=2, value=user.username)
            ws_main.cell(row=row, column=3, value=user.email)
            ws_main.cell(row=row, column=4, value=user.phone or '')
            ws_main.cell(row=row, column=5, value=role_labels.get(user.role, user.role))
            ws_main.cell(row=row, column=6, value=status_labels.get(user.status, user.status))
            ws_main.cell(row=row, column=7, value=user.governorate or '')
            ws_main.cell(row=row, column=8, value=user.delegation or '')
            ws_main.cell(row=row, column=9, value=user.farm_name or '')
            ws_main.cell(row=row, column=10, value=user.farm_address or '')
            ws_main.cell(row=row, column=11, value=user.farm_lat or '')
            ws_main.cell(row=row, column=12, value=user.farm_lng or '')
            
            # Date inscription
            if user.date_joined:
                ws_main.cell(row=row, column=13, value=user.date_joined.strftime('%d/%m/%Y %H:%M'))
            else:
                ws_main.cell(row=row, column=13, value='')
            
            # Dernière connexion
            if user.last_login:
                ws_main.cell(row=row, column=14, value=user.last_login.strftime('%d/%m/%Y %H:%M'))
            else:
                ws_main.cell(row=row, column=14, value='Jamais')
            
            # Ajouter une couleur de fond selon le statut
            status_fills = {
                'ACTIVE': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),  # Vert clair
                'SUSPENDED': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),  # Orange clair
                'PENDING': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),  # Bleu clair
                'BLOCKED': PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),  # Rouge clair
            }
            
            if user.status in status_fills:
                for col in range(1, 15):
                    cell = ws_main.cell(row=row, column=col)
                    cell.fill = status_fills[user.status]
            
            row += 1
        
        # === FEUILLE 2 : STATISTIQUES PAR RÔLE ===
        ws_stats = wb.create_sheet("Statistiques")
        
        ws_stats.merge_cells('A1:E1')
        ws_stats['A1'] = "STATISTIQUES DES UTILISATEURS PAR RÔLE"
        ws_stats['A1'].font = Font(name='Arial', size=14, bold=True, color='2E7D32')
        ws_stats['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Calculer les statistiques par rôle
        stats_headers = ["ROLE", "TOTAL", "ACTIFS", "SUSPENDUS", "EN ATTENTE", "BLOQUÉS"]
        for col_idx, header in enumerate(stats_headers, start=1):
            cell = ws_stats.cell(row=3, column=col_idx, value=header)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        
        # Récupérer les statistiques depuis la base de données
        stats_row = 4
        roles = ['FARMER', 'BENEFICIARY', 'DUAL', 'ADMIN', 'SUPERADMIN']
        
        for role in roles:
            role_users = users_qs.filter(role=role)
            if role_users.exists():
                ws_stats.cell(row=stats_row, column=1, value=role_labels.get(role, role))
                ws_stats.cell(row=stats_row, column=2, value=role_users.count())
                ws_stats.cell(row=stats_row, column=3, value=role_users.filter(status='ACTIVE').count())
                ws_stats.cell(row=stats_row, column=4, value=role_users.filter(status='SUSPENDED').count())
                ws_stats.cell(row=stats_row, column=5, value=role_users.filter(status='PENDING').count())
                ws_stats.cell(row=stats_row, column=6, value=role_users.filter(status='BLOCKED').count())
                stats_row += 1
        
        # Total général
        ws_stats.cell(row=stats_row, column=1, value="TOTAL")
        ws_stats.cell(row=stats_row, column=1).font = Font(name='Arial', size=10, bold=True)
        ws_stats.cell(row=stats_row, column=2, value=users_qs.count())
        ws_stats.cell(row=stats_row, column=2).font = Font(name='Arial', size=10, bold=True)
        ws_stats.cell(row=stats_row, column=3, value=users_qs.filter(status='ACTIVE').count())
        ws_stats.cell(row=stats_row, column=3).font = Font(name='Arial', size=10, bold=True)
        ws_stats.cell(row=stats_row, column=4, value=users_qs.filter(status='SUSPENDED').count())
        ws_stats.cell(row=stats_row, column=4).font = Font(name='Arial', size=10, bold=True)
        
        # === FEUILLE 3 : RÉPARTITION GÉOGRAPHIQUE ===
        ws_geo = wb.create_sheet("Géographie")
        
        ws_geo.merge_cells('A1:D1')
        ws_geo['A1'] = "RÉPARTITION GÉOGRAPHIQUE DES UTILISATEURS"
        ws_geo['A1'].font = Font(name='Arial', size=14, bold=True, color='2E7D32')
        ws_geo['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        geo_headers = ["GOUVERNORAT", "NOMBRE D'UTILISATEURS", "FARMERS", "BÉNÉFICIAIRES", "DUALS"]
        for col_idx, header in enumerate(geo_headers, start=1):
            cell = ws_geo.cell(row=3, column=col_idx, value=header)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        
        # Récupérer les statistiques géographiques
        governorate_stats = users_qs.exclude(governorate__isnull=True).exclude(governorate='')\
            .values('governorate')\
            .annotate(
                total=Count('id'),
                farmers=Count('id', filter=Q(role='FARMER')),
                beneficiaries=Count('id', filter=Q(role='BENEFICIARY')),
                duals=Count('id', filter=Q(role='DUAL'))
            )\
            .order_by('-total')
        
        geo_row = 4
        for stat in governorate_stats:
            ws_geo.cell(row=geo_row, column=1, value=stat['governorate'])
            ws_geo.cell(row=geo_row, column=2, value=stat['total'])
            ws_geo.cell(row=geo_row, column=3, value=stat['farmers'])
            ws_geo.cell(row=geo_row, column=4, value=stat['beneficiaries'])
            ws_geo.cell(row=geo_row, column=5, value=stat['duals'])
            geo_row += 1
        
        # Total par gouvernorat
        ws_geo.cell(row=geo_row, column=1, value="TOTAL")
        ws_geo.cell(row=geo_row, column=1).font = Font(name='Arial', size=10, bold=True)
        ws_geo.cell(row=geo_row, column=2, value=users_qs.exclude(governorate__isnull=True).exclude(governorate='').count())
        ws_geo.cell(row=geo_row, column=2).font = Font(name='Arial', size=10, bold=True)
        
        # === FEUILLE 4 : ACTIVITÉ RÉCENTE ===
        ws_activity = wb.create_sheet("Activité")
        
        ws_activity.merge_cells('A1:C1')
        ws_activity['A1'] = "ACTIVITÉ RÉCENTE DES UTILISATEURS"
        ws_activity['A1'].font = Font(name='Arial', size=14, bold=True, color='2E7D32')
        ws_activity['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Utilisateurs récemment inscrits (30 derniers jours)
        from datetime import timedelta
        from django.utils import timezone
        
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_users = users_qs.filter(date_joined__gte=thirty_days_ago).order_by('-date_joined')
        
        ws_activity['A3'] = f"Nouveaux utilisateurs (30 derniers jours): {recent_users.count()}"
        ws_activity.merge_cells('A3:C3')
        ws_activity['A3'].font = Font(name='Arial', size=11, bold=True)
        
        activity_headers = ["Nom d'utilisateur", "Rôle", "Date d'inscription", "Statut"]
        for col_idx, header in enumerate(activity_headers, start=1):
            cell = ws_activity.cell(row=5, column=col_idx, value=header)
            cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
        
        activity_row = 6
        for user in recent_users[:20]:  # Limiter à 20 pour éviter un fichier trop lourd
            ws_activity.cell(row=activity_row, column=1, value=user.username)
            ws_activity.cell(row=activity_row, column=2, value=role_labels.get(user.role, user.role))
            ws_activity.cell(row=activity_row, column=3, value=user.date_joined.strftime('%d/%m/%Y %H:%M'))
            ws_activity.cell(row=activity_row, column=4, value=status_labels.get(user.status, user.status))
            activity_row += 1
        
        # Appliquer les styles
        _apply_excel_styles(ws_main)
        _apply_excel_styles(ws_stats)
        _apply_excel_styles(ws_geo)
        _apply_excel_styles(ws_activity)
        
        # Créer la réponse HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"utilisateurs_agricycle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Sauvegarder le workbook dans la réponse
        wb.save(response)
        return response
        
    except Exception as e:
        print(f"Erreur lors de l'export Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        from rest_framework.response import Response
        from rest_framework import status
        return Response(
            {"detail": f"Erreur lors de la génération du fichier Excel: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
@api_view(["POST"])
@permission_classes([AllowAny])
def register(request):
    username = request.data.get("username", "").strip()
    email = request.data.get("email", "").strip().lower()
    password = request.data.get("password", "")
    role = request.data.get("role", User.Role.FARMER)
    phone = request.data.get("phone", "").strip()
    city = request.data.get("city", "").strip()

    errors = {}

    # Validation username
    if not username:
        errors["username"] = ["Le nom d'utilisateur est requis"]
    elif len(username) < 3:
        errors["username"] = ["Le nom d'utilisateur doit contenir au moins 3 caractères"]

    # Validation email
    if not email:
        errors["email"] = ["L'email est requis"]
    else:
        try:
            validate_email(email)
        except ValidationError:
            errors["email"] = ["Veuillez entrer une adresse email valide"]

    # Validation password
    if not password:
        errors["password"] = ["Le mot de passe est requis"]
    else:
        try:
            validate_password(password)
        except ValidationError as e:
            errors["password"] = list(e.messages)

    # Validation city
    if not city:
        errors["city"] = ["La ville est requise"]

    # Email unique
    if email and User.objects.filter(email=email).exists():
        errors["email"] = ["Cette adresse email est déjà utilisée"]

    # Role validation
    valid_roles = [choice[0] for choice in User.Role.choices]
    if role not in valid_roles:
        errors["role"] = [f"Rôle invalide. Choisissez parmi: {', '.join(valid_roles)}"]

    # Phone unique (optionnel)
    if phone and User.objects.filter(phone=phone).exists():
        errors["phone"] = ["Ce numéro de téléphone est déjà utilisé"]

    if errors:
        return Response({
            "detail": "Veuillez corriger les erreurs ci-dessous",
            "errors": errors
        }, status=status.HTTP_400_BAD_REQUEST)

    # Protection admin
    if role in [User.Role.ADMIN, User.Role.SUPERADMIN]:
        if not request.user.is_authenticated or not request.user.is_admin_role():
            return Response(
                {"detail": "Vous n'avez pas l'autorisation de créer un compte administrateur"},
                status=status.HTTP_403_FORBIDDEN
            )

    try:
        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            role=role,
            phone=phone if phone else None,
            city=city
        )

        return Response({
            "message": "Inscription réussie ! Vous pouvez maintenant vous connecter",
            "user": {
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "role": user.role,
                "city": user.city,
            }
        }, status=status.HTTP_201_CREATED)

    except Exception as e:
        return Response(
            {"detail": f"Erreur lors de l'inscription: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ---------- Auth ----------
@api_view(["POST"])
@permission_classes([AllowAny])
def token_login(request):
    identifier = (request.data.get("identifier") or "").strip()  # email ou téléphone
    password = request.data.get("password") or ""

    if not identifier or not password:
        return Response({"detail": "Les champs identifiant et mot de passe sont requis."},
                        status=status.HTTP_400_BAD_REQUEST)

    # Vérifier si l'identifiant est un email ou un numéro de téléphone
    if "@" in identifier:
        # Chercher par email
        user = User.objects.filter(email__iexact=identifier).first()
    else:
        # Chercher par numéro de téléphone
        user = User.objects.filter(phone=identifier).first()

    if not user:
        return Response({"detail": "Identifiants invalides"}, status=status.HTTP_401_UNAUTHORIZED)

    # Authentifier l'utilisateur
    user_auth = authenticate(request, username=user.username, password=password)
    if not user_auth:
        return Response({"detail": "Identifiants invalides"}, status=status.HTTP_401_UNAUTHORIZED)

    # Créer le token de rafraîchissement et d'accès
    refresh = RefreshToken.for_user(user_auth)

    # Retourner les tokens
    return Response({
        "access": str(refresh.access_token),
        "refresh": str(refresh),
    })



def hash_code(code: str) -> str:
    """Hash un code OTP pour le stockage sécurisé"""
    return hashlib.sha256(code.encode()).hexdigest()
@api_view(["POST"])
@permission_classes([AllowAny])
def reset_password_request(request):
    """
    Demande de réinitialisation avec OTP via Resend
    """
    email = request.data.get("email", "").strip().lower()
    
    # Log pour debug
    print(f"📧 Demande reset password pour: {email}")
    print(f"📦 Données reçues: {request.data}")
    
    # Validation
    if not email:
        print("❌ Erreur: Email vide")
        return Response({"detail": "L'email est obligatoire"}, status=status.HTTP_400_BAD_REQUEST)
    
    try:
        validate_email(email)
    except ValidationError:
        print(f"❌ Erreur: Email invalide: {email}")
        return Response({"detail": "Email invalide"}, status=status.HTTP_400_BAD_REQUEST)
    
    user = User.objects.filter(email=email).first()
    if not user:
        print(f"⚠️  Aucun utilisateur trouvé pour: {email}")
        # Pour la sécurité, on ne dit pas si l'email existe
        return Response({"message": "Si cet email existe, un code OTP sera envoyé"})
    
    print(f"✅ Utilisateur trouvé: {user.username}")
    
    # Générer OTP
    import random
    code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    print(f"🔑 Code OTP généré: {code}")
    
    # Sauvegarder OTP
    try:
        otp_obj = OTPCode.objects.create(
            email=email,
            code=code,
            code_hash=hash_code(code),
            expires_at=timezone.now() + timedelta(minutes=15),
            is_used=False,
            purpose=OTPCode.Purpose.RESET_PASSWORD,
            attempts=0,
        )
        print(f"💾 OTP sauvegardé en base: ID={otp_obj.id}")
    except Exception as e:
        print(f"❌ Erreur sauvegarde OTP: {e}")
        return Response({"detail": "Erreur interne"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    # Envoyer avec Resend
    try:
        import resend
        from django.conf import settings
        
        print(f"🔧 Configuration Resend: API_KEY={settings.RESEND_API_KEY[:10]}...")
        print(f"🔧 Email from: {settings.DEFAULT_FROM_EMAIL}")
        
        # Vérifier que Resend est configuré
        if not settings.RESEND_API_KEY:
            print("⚠️  RESEND_API_KEY non configurée")
            raise ValueError("RESEND_API_KEY manquante")
        
        # Préparer le contenu HTML
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
                .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                .header {{ background: #10b981; color: white; padding: 20px; text-align: center; }}
                .otp-code {{ 
                    font-size: 32px; 
                    font-weight: bold; 
                    text-align: center; 
                    margin: 30px 0;
                    letter-spacing: 5px;
                    color: #059669;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🔐 Réinitialisation de mot de passe</h1>
                </div>
                <p>Bonjour,</p>
                <p>Vous avez demandé à réinitialiser votre mot de passe AgriCycle.</p>
                <p>Votre code OTP est :</p>
                <div class="otp-code">{code}</div>
                <p>Ce code est valable pendant <strong>15 minutes</strong>.</p>
                <p>Si vous n'avez pas fait cette demande, ignorez cet email.</p>
                <br>
                <p>Cordialement,<br>L'équipe AgriCycle</p>
            </div>
        </body>
        </html>
        """
        
        # Envoyer avec Resend
        params = {
            "from": settings.DEFAULT_FROM_EMAIL,
            "to": [email],
            "subject": "🔐 Code OTP pour réinitialisation - AgriCycle",
            "html": html_content,
        }
        
        print(f"📤 Envoi email à {email}...")
        response = resend.Emails.send(params)
        print(f"✅ Email envoyé! Response: {response}")
        
        # Toujours retourner le même message pour la sécurité
        return Response({
            "message": "Si cet email existe, un code OTP sera envoyé",
            "debug": "email_envoye"  # Pour debug seulement
        })
        
    except Exception as e:
        print(f"❌ Erreur Resend: {str(e)}")
        print(f"🔍 Type d'erreur: {type(e).__name__}")
        import traceback
        traceback.print_exc()
        
        # Fallback: afficher l'OTP dans la console pour le développement
        print(f"🚨 EN MODE DEVELOPPEMENT - OTP pour {email}: {code}")
        print(f"🚨 Pour le test, utilisez ce code: {code}")
        
        # Même réponse pour la sécurité
        return Response({
            "message": "Si cet email existe, un code OTP sera envoyé",
            "debug": f"mode_dev_otp: {code}"  # Pour debug seulement
        })


@api_view(["POST"])
@permission_classes([AllowAny])
def reset_password_confirm(request):
    """
    Confirmation avec OTP et nouveau mot de passe
    """
    email = request.data.get("email", "").strip().lower()
    otp = request.data.get("otp", "")
    new_password = request.data.get("new_password", "")

    # Validation des champs
    if not all([email, otp, new_password]):
        return Response({"detail": "Tous les champs sont requis"}, status=status.HTTP_400_BAD_REQUEST)
    
    # Récupérer le dernier OTP non utilisé
    otp_obj = OTPCode.objects.filter(
        email=email, 
        purpose=OTPCode.Purpose.RESET_PASSWORD, 
        is_used=False
    ).order_by("-created_at").first()
    
    if not otp_obj:
        return Response({"detail": "Code OTP invalide ou expiré"}, status=status.HTTP_400_BAD_REQUEST)
    
    # Vérifier l'expiration
    if otp_obj.expires_at < timezone.now():
        otp_obj.is_used = True
        otp_obj.save()
        return Response({"detail": "Code OTP expiré"}, status=status.HTTP_400_BAD_REQUEST)
    
    # Vérifier le code
    if otp_obj.code_hash != hash_code(otp):
        # Incrémenter les tentatives
        otp_obj.attempts += 1
        otp_obj.save()
        
        if otp_obj.attempts >= 3:
            otp_obj.is_used = True
            otp_obj.save()
            return Response({"detail": "Trop de tentatives erronées"}, status=status.HTTP_400_BAD_REQUEST)
        
        return Response({"detail": "Code OTP incorrect"}, status=status.HTTP_400_BAD_REQUEST)
    
    # Valider le mot de passe
    try:
        validate_password(new_password)
    except ValidationError as e:
        return Response({"detail": list(e.messages)}, status=status.HTTP_400_BAD_REQUEST)
    
    # Récupérer l'utilisateur
    user = User.objects.filter(email=email).first()
    if not user:
        return Response({"detail": "Utilisateur introuvable"}, status=status.HTTP_404_NOT_FOUND)
    
    # Mettre à jour le mot de passe
    user.set_password(new_password)
    user.save()
    
    # Marquer l'OTP comme utilisé
    otp_obj.is_used = True
    otp_obj.save()
    
    return Response({"message": "Mot de passe modifié avec succès"})

# ---------- Health Check ----------
@api_view(["GET"])
@permission_classes([AllowAny])
def api_health(request):
    return Response({
        "status": "ok",
        "service": "AgriCycle API",
    })


# ---------- Stats Views ----------
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def stats_farmer(request):
    u = request.user
    if u.role not in {User.Role.FARMER, User.Role.DUAL}:
        return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

    qs = Declaration.objects.filter(owner=u)
    total_qty = qs.aggregate(total=Sum("quantity"))["total"] or 0
    by_status = qs.values("status").annotate(c=Count("id")).order_by("status")
    return Response({
        "total_quantity": total_qty,
        "declarations_count": qs.count(),
        "by_status": list(by_status),
        "farm": {"farm_lat": u.farm_lat, "farm_lng": u.farm_lng, "farm_address": u.farm_address, "farm_name": u.farm_name},
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def stats_beneficiary(request):
    u = request.user
    if u.role not in {User.Role.BENEFICIARY, User.Role.DUAL}:
        return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

    oqs = Order.objects.filter(buyer=u).exclude(status=Order.Status.CART)
    spent = oqs.aggregate(s=Sum("total_amount"))["s"] or 0
    by_status = oqs.values("status").annotate(c=Count("id")).order_by("status")

    # products stats
    products_count = Product.objects.filter(is_active=True).count()
    by_cat = Product.objects.filter(is_active=True).values("category").annotate(c=Count("id")).order_by("-c")

    return Response({
        "orders_count": oqs.count(),
        "total_spent": spent,
        "orders_by_status": list(by_status),
        "products_count": products_count,
        "products_by_category": list(by_cat),
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def stats_admin(request):
    """API pour les statistiques du dashboard admin - TOUS les types"""
    try:
        # Vérification admin
        if not _admin_only(request.user):
            return Response({"detail": "Forbidden - Admin uniquement"}, status=status.HTTP_403_FORBIDDEN)
        
        print(f"✅ Accès autorisé pour {request.user.username}")
        
        # 1. Compter les agriculteurs actifs (FARMER seulement)
        farmers_count = User.objects.filter(
            role='FARMER',
            is_active=True
        ).count()

        # 2. Compter les bénéficiaires (BENEFICIARY seulement)
        beneficiaries_count = User.objects.filter(
            role='BENEFICIARY',
            is_active=True
        ).count()

        # 3. Compter les utilisateurs DUAL
        dual_count = User.objects.filter(
            role='DUAL',
            is_active=True
        ).count()

        # 4. TOTAL des utilisateurs actifs
        total_active_users = User.objects.filter(is_active=True).count()

        # 5. TOTAL des déclarations
        declarations_count = Declaration.objects.count()

        # 6. TOTAL des commandes (hors "abandoned" et "CART")
        orders_count = Order.objects.exclude(
            Q(status="abandoned") | Q(status="CART")
        ).count()

        # 7. Revenu total (commandes "completed", "delivered", "paid", "CONFIRMED")
        revenue = Order.objects.filter(
            Q(status="completed") | Q(status="delivered") | 
            Q(status="paid") | Q(status="CONFIRMED")
        ).aggregate(
            total=Sum('total_amount')
        )['total'] or 0

        # 8. Nombre de produits actifs
        products_count = Product.objects.filter(is_active=True).count()

        # 9. Données pour la carte (tous les utilisateurs avec coordonnées)
        users = User.objects.filter(
            Q(role='FARMER') | Q(role='BENEFICIARY') | Q(role='DUAL'),
            farm_lat__isnull=False,
            farm_lng__isnull=False
        )
        
        users_points = []
        for user in users:
            try:
                users_points.append({
                    "id": user.id,
                    "username": user.username,
                    "governorate": user.governorate or "",
                    "delegation": user.delegation or "",
                    "farm_lat": float(user.farm_lat) if user.farm_lat else 0,
                    "farm_lng": float(user.farm_lng) if user.farm_lng else 0,
                    "farm_name": user.farm_name or "",
                    "role": user.role,
                    "farm_address": user.farm_address or ""
                })
            except (ValueError, TypeError) as e:
                continue

        # 10. Statistiques par statut
        decl_by_status = list(Declaration.objects.values("status").annotate(
            c=Count("id")
        ).order_by("status"))
        
        orders_by_status = list(Order.objects.values("status").annotate(
            c=Count("id")
        ).order_by("status"))

        # Debug logging
        print(f"📊 Stats calculées:")
        print(f"  Farmers (FARMER): {farmers_count}")
        print(f"  Beneficiaries (BENEFICIARY): {beneficiaries_count}")
        print(f"  Dual (DUAL): {dual_count}")
        print(f"  Total users: {total_active_users}")
        print(f"  Declarations: {declarations_count}")
        print(f"  Orders: {orders_count}")
        print(f"  Revenue: {revenue}")
        print(f"  Products: {products_count}")
        print(f"  Map points: {len(users_points)}")

        response_data = {
            # Stats pour les cartes du dashboard
            "farmers": farmers_count,           # Seulement FARMER
            "declarations": declarations_count,
            "orders": orders_count,
            "revenue": float(revenue),
            "beneficiaries": beneficiaries_count,  # Seulement BENEFICIARY
            "products": products_count,
            "dual_users": dual_count,           # Seulement DUAL
            
            # Totaux additionnels
            "total_farmers_dual": farmers_count + dual_count,  # Agriculteurs totaux (FARMER + DUAL)
            "total_beneficiaries_dual": beneficiaries_count + dual_count,  # Bénéficiaires totaux
            "total_users": total_active_users,
            "total_orders_amount": float(revenue),
            
            # Données pour la carte (tous types)
            "users_points": users_points,  # Renommer pour plus de clarté
            
            # Stats détaillées
            "declarations_by_status": decl_by_status,
            "orders_by_status": orders_by_status,
            
            # Timestamp
            "last_updated": timezone.now().isoformat()
        }

        return Response(response_data)

    except Exception as e:
        print(f"❌ Erreur dans stats_admin: {e}")
        traceback.print_exc()
        
        return Response({
            "error": str(e),
            "farmers": 0,
            "declarations": 0,
            "orders": 0,
            "revenue": 0,
            "beneficiaries": 0,
            "products": 0,
            "dual_users": 0,
            "users_points": [],
            "declarations_by_status": [],
            "orders_by_status": []
        }, status=500)
        
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def admin_farmers_map(request):
    """API pour les données de la carte - TOUS les utilisateurs avec coordonnées"""
    try:
        # Vérification admin
        if not _admin_only(request.user):
            return Response({"detail": "Forbidden - Admin uniquement"}, status=status.HTTP_403_FORBIDDEN)

        print(f"🗺️ Requête carte reçue de {request.user.username}")
        
        # Récupérer TOUS les utilisateurs avec coordonnées (FARMER, BENEFICIARY, DUAL)
        users = User.objects.filter(
            Q(role='FARMER') | Q(role='BENEFICIARY') | Q(role='DUAL'),
            farm_lat__isnull=False,
            farm_lng__isnull=False
        ).select_related()
        
        print(f"📍 Requête SQL: {users.query}")
        print(f"📍 Nombre d'utilisateurs trouvés: {users.count()}")
        
        users_points = []
        for user in users:
            try:
                users_points.append({
                    "id": user.id,
                    "username": user.username,
                    "governorate": user.governorate or "",
                    "delegation": user.delegation or "",
                    "farm_lat": float(user.farm_lat) if user.farm_lat else 0,
                    "farm_lng": float(user.farm_lng) if user.farm_lng else 0,
                    "farm_name": user.farm_name or "",
                    "role": user.role,
                    "farm_address": user.farm_address or "",
                    "email": user.email or "",
                    "phone": user.phone or "",
                    "status": user.status if hasattr(user, 'status') else "ACTIVE",
                    "is_active": user.is_active
                })
                print(f"  ✅ {user.username}: lat={user.farm_lat}, lng={user.farm_lng}, role={user.role}")
            except (ValueError, TypeError) as e:
                print(f"  ❌ Erreur conversion pour {user.username}: {e}")
                continue
        
        print(f"📍 Carte: {len(users_points)} utilisateurs chargés")
        print(f"  - FARMER: {len([u for u in users_points if u['role'] == 'FARMER'])}")
        print(f"  - BENEFICIARY: {len([u for u in users_points if u['role'] == 'BENEFICIARY'])}")
        print(f"  - DUAL: {len([u for u in users_points if u['role'] == 'DUAL'])}")
        
        # Log les 3 premiers pour debug
        if users_points:
            print("📍 Exemple données retournées:", json.dumps(users_points[:3], indent=2, ensure_ascii=False))
        
        return Response(users_points)
        
    except Exception as e:
        print(f"❌ Erreur dans admin_farmers_map: {e}")
        import traceback
        traceback.print_exc()
        return Response({"error": str(e), "users_points": []}, status=500)
# ---------- Exports ----------
@api_view(["GET"])
def export_orders_report_xlsx(request):
    if not _admin_only(request.user):
        return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)
    
    status_filter = request.query_params.get("status")
    date_from = request.query_params.get("from")  # YYYY-MM-DD
    date_to = request.query_params.get("to")      # YYYY-MM-DD
    
    qs = Order.objects.all().prefetch_related("items", "items__product").order_by("-created_at")
    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)
    
    revenue_qs = qs.exclude(status__in=[Order.Status.CART, Order.Status.CANCELLED])
    
    by_status = qs.values("status").annotate(
        orders_count=Count("id"), 
        amount_sum=Sum("total_amount")
    ).order_by("status")
    
    totals = qs.aggregate(orders_count=Count("id"), amount_sum=Sum("total_amount"))
    revenue_totals = revenue_qs.aggregate(
        revenue_orders_count=Count("id"), 
        revenue_sum=Sum("total_amount")
    )
    
    line_total = ExpressionWrapper(
        F("qty") * F("unit_price"), 
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )
    
    top_products = (
        OrderItem.objects.filter(order__in=revenue_qs)
        .values("product_id", "product__name")
        .annotate(qty_sum=Sum("qty"), revenue_sum=Sum(line_total))
        .order_by("-revenue_sum")[:20]
    )
    
    top_buyers = (
        revenue_qs.values("buyer_id", "buyer__username")
        .annotate(orders_count=Count("id"), spent_sum=Sum("total_amount"))
        .order_by("-spent_sum")[:20]
    )
    
    top_regions = (
        revenue_qs.values("buyer__governorate")
        .annotate(orders_count=Count("id"), revenue_sum=Sum("total_amount"))
        .order_by("-revenue_sum")[:20]
    )
    
    # Trend monthly revenue
    monthly = defaultdict(float)
    for o in revenue_qs:
        key = o.created_at.strftime("%Y-%m")
        monthly[key] += float(o.total_amount)
    trend_rows = sorted(monthly.items())  # (YYYY-MM, revenue)
    
    # Création du workbook
    wb = Workbook()
    
    # Supprimer la feuille par défaut
    if 'Sheet' in wb.sheetnames:
        default_ws = wb['Sheet']
        wb.remove(default_ws)
    
    # === FEUILLE 1 : RÉSUMÉ GÉNÉRAL ===
    ws_summary = wb.create_sheet("Résumé")
    
    # Titre principal
    ws_summary.merge_cells('A1:E1')
    title_cell = ws_summary['A1']
    title_cell.value = "RAPPORT DES COMMANDES - AGRI CYCLE"
    title_cell.font = Font(name='Arial', size=16, bold=True, color='2E7D32')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Informations de génération
    ws_summary['A2'] = f"Généré le : {datetime.now().strftime('%d/%m/%Y à %H:%M')} | Total: {qs.count()} commandes"
    ws_summary.merge_cells('A2:E2')
    ws_summary['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws_summary['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Filtres appliqués
    ws_summary['A3'] = f"Filtres: Statut={status_filter or 'Tous'} | Période: {date_from or 'Début'} au {date_to or 'Aujourd hui'}"
    ws_summary.merge_cells('A3:E3')
    ws_summary['A3'].font = Font(name='Arial', size=9, italic=True, color='666666')
    ws_summary['A3'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Section Totaux Généraux
    ws_summary['A5'] = "📊 TOTAUX GÉNÉRAUX"
    ws_summary['A5'].font = Font(name='Arial', size=12, bold=True, color='2E7D32')
    ws_summary['A5'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    
    summary_data = [
        ["Description", "Valeur"],
        ["Toutes les commandes", totals["orders_count"] or 0],
        ["Montant total", f"{totals['amount_sum'] or 0:,.2f} DT"],
        ["Commandes validées (hors paniers)", revenue_totals["revenue_orders_count"] or 0],
        ["Revenu généré", f"{revenue_totals['revenue_sum'] or 0:,.2f} DT"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 6:  # En-tête
                cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            elif row_idx % 2 == 0:  # Lignes paires
                cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    # Section Répartition par Statut
    ws_summary['A13'] = "📈 RÉPARTITION PAR STATUT"
    ws_summary['A13'].font = Font(name='Arial', size=12, bold=True, color='2E7D32')
    ws_summary['A13'].fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    
    status_headers = ["Statut", "Nombre de Commandes", "Montant Total (DT)"]
    for col_idx, header in enumerate(status_headers, start=1):
        cell = ws_summary.cell(row=14, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    
    status_row = 15
    for r in by_status:
        ws_summary.cell(row=status_row, column=1, value=r["status"])
        ws_summary.cell(row=status_row, column=2, value=r["orders_count"])
        ws_summary.cell(row=status_row, column=3, value=f"{r['amount_sum'] or 0:,.2f} DT")
        status_row += 1
    
    # === FEUILLE 2 : TOP PRODUITS ===
    ws_products = wb.create_sheet("Top Produits")
    
    ws_products.merge_cells('A1:D1')
    ws_products['A1'] = "🏆 TOP PRODUITS PAR REVENUS"
    ws_products['A1'].font = Font(name='Arial', size=14, bold=True, color='2E7D32')
    ws_products['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_products['A2'] = f"Total: {len(top_products)} produits | Généré le: {datetime.now().strftime('%d/%m/%Y')}"
    ws_products.merge_cells('A2:D2')
    ws_products['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws_products['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    product_headers = ["ID Produit", "Nom du Produit", "Quantité Vendue", "Revenu Généré (DT)"]
    for col_idx, header in enumerate(product_headers, start=1):
        cell = ws_products.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    
    product_row = 5
    for p in top_products:
        ws_products.cell(row=product_row, column=1, value=p["product_id"])
        ws_products.cell(row=product_row, column=2, value=p["product__name"])
        ws_products.cell(row=product_row, column=3, value=p["qty_sum"] or 0)
        ws_products.cell(row=product_row, column=4, value=f"{p['revenue_sum'] or 0:,.2f} DT")
        
        # Couleur pour le top 3
        if product_row <= 7:
            fill_color = 'FFF3E0' if product_row == 5 else 'E8F5E9' if product_row == 6 else 'E3F2FD'
            for col in range(1, 5):
                ws_products.cell(row=product_row, column=col).fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type='solid'
                )
        
        product_row += 1
    
    # === FEUILLE 3 : TOP ACHETEURS ===
    ws_buyers = wb.create_sheet("Top Acheteurs")
    
    ws_buyers.merge_cells('A1:E1')
    ws_buyers['A1'] = "👥 TOP ACHETEURS"
    ws_buyers['A1'].font = Font(name='Arial', size=14, bold=True, color='2196F3')
    ws_buyers['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    buyer_headers = ["ID Acheteur", "Nom d'utilisateur", "Nombre de Commandes", "Total Dépensé (DT)", "Moyenne par Commande"]
    for col_idx, header in enumerate(buyer_headers, start=1):
        cell = ws_buyers.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
    
    buyer_row = 4
    for b in top_buyers:
        ws_buyers.cell(row=buyer_row, column=1, value=b["buyer_id"])
        ws_buyers.cell(row=buyer_row, column=2, value=b["buyer__username"])
        ws_buyers.cell(row=buyer_row, column=3, value=b["orders_count"] or 0)
        spent = b["spent_sum"] or 0
        ws_buyers.cell(row=buyer_row, column=4, value=f"{spent:,.2f} DT")
        avg = spent / b["orders_count"] if b["orders_count"] > 0 else 0
        ws_buyers.cell(row=buyer_row, column=5, value=f"{avg:,.2f} DT")
        
        buyer_row += 1
    
    # === FEUILLE 4 : TOP RÉGIONS ===
    ws_regions = wb.create_sheet("Top Régions")
    
    ws_regions.merge_cells('A1:D1')
    ws_regions['A1'] = "📍 TOP RÉGIONS PAR REVENUS"
    ws_regions['A1'].font = Font(name='Arial', size=14, bold=True, color='9C27B0')
    ws_regions['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    region_headers = ["Gouvernorat", "Nombre de Commandes", "Revenu Total (DT)", "Moyenne par Commande"]
    for col_idx, header in enumerate(region_headers, start=1):
        cell = ws_regions.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
    
    region_row = 4
    for r in top_regions:
        ws_regions.cell(row=region_row, column=1, value=r["buyer__governorate"] or "Non spécifié")
        ws_regions.cell(row=region_row, column=2, value=r["orders_count"] or 0)
        revenue = r["revenue_sum"] or 0
        ws_regions.cell(row=region_row, column=3, value=f"{revenue:,.2f} DT")
        avg = revenue / r["orders_count"] if r["orders_count"] > 0 else 0
        ws_regions.cell(row=region_row, column=4, value=f"{avg:,.2f} DT")
        
        region_row += 1
    
    # === FEUILLE 5 : TENDANCES MENSUELLES ===
    ws_trend = wb.create_sheet("Tendances Mensuelles")
    
    ws_trend.merge_cells('A1:C1')
    ws_trend['A1'] = "📅 TENDANCES MENSUELLES"
    ws_trend['A1'].font = Font(name='Arial', size=14, bold=True, color='FF9800')
    ws_trend['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    trend_headers = ["Mois (YYYY-MM)", "Revenu (DT)", "Nombre de Commandes"]
    for col_idx, header in enumerate(trend_headers, start=1):
        cell = ws_trend.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
    
    trend_row = 4
    for m, rev in trend_rows[-12:]:  # Derniers 12 mois
        ws_trend.cell(row=trend_row, column=1, value=m)
        ws_trend.cell(row=trend_row, column=2, value=f"{rev:,.2f} DT")
        
        # Compter les commandes pour ce mois
        month_orders = revenue_qs.filter(created_at__year=int(m[:4]), created_at__month=int(m[5:7])).count()
        ws_trend.cell(row=trend_row, column=3, value=month_orders)
        
        trend_row += 1
    
    # === FEUILLE 6 : LISTE DES COMMANDES ===
    ws_orders = wb.create_sheet("Liste des Commandes")
    
    ws_orders.merge_cells('A1:G1')
    ws_orders['A1'] = "📋 LISTE DÉTAILLÉE DES COMMANDES"
    ws_orders['A1'].font = Font(name='Arial', size=16, bold=True, color='2E7D32')
    ws_orders['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws_orders['A2'] = f"Total: {qs.count()} commandes | Généré le: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_orders.merge_cells('A2:G2')
    ws_orders['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws_orders['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    order_headers = [
        "ID Commande", "ID Acheteur", "Nom Acheteur", 
        "Date Création", "Statut", "Montant Total (DT)", "Nombre d'Articles"
    ]
    for col_idx, header in enumerate(order_headers, start=1):
        cell = ws_orders.cell(row=4, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    
    order_row = 5
    for o in qs:
        ws_orders.cell(row=order_row, column=1, value=o.id)
        ws_orders.cell(row=order_row, column=2, value=o.buyer_id)
        ws_orders.cell(row=order_row, column=3, value=o.buyer.username if o.buyer else "N/A")
        ws_orders.cell(row=order_row, column=4, value=o.created_at.strftime('%d/%m/%Y %H:%M'))
        ws_orders.cell(row=order_row, column=5, value=o.status)
        ws_orders.cell(row=order_row, column=6, value=f"{o.total_amount:,.2f} DT")
        ws_orders.cell(row=order_row, column=7, value=o.items.count())
        
        # Couleur selon le statut
        status_colors = {
            'completed': 'E8F5E9',  # Vert clair
            'delivered': 'E8F5E9',
            'paid': 'E8F5E9',
            'CONFIRMED': 'FFF3E0',  # Orange clair
            'CART': 'E3F2FD',  # Bleu clair
            'CANCELLED': 'FFEBEE',  # Rouge clair
            'abandoned': 'FFEBEE',
        }
        
        if o.status in status_colors:
            for col in range(1, 8):
                ws_orders.cell(row=order_row, column=col).fill = PatternFill(
                    start_color=status_colors[o.status], 
                    end_color=status_colors[o.status], 
                    fill_type='solid'
                )
        
        order_row += 1
    
    # === FEUILLE 7 : DÉTAILS DES ARTICLES ===
    ws_items = wb.create_sheet("Détails Articles")
    
    ws_items.merge_cells('A1:H1')
    ws_items['A1'] = "📦 DÉTAILS DES ARTICLES DE COMMANDE"
    ws_items['A1'].font = Font(name='Arial', size=16, bold=True, color='2E7D32')
    ws_items['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    item_headers = [
        "ID Commande", "ID Article", "ID Produit", "Nom Produit", 
        "Quantité", "Prix Unitaire (DT)", "Total Ligne (DT)", "Statut Commande"
    ]
    for col_idx, header in enumerate(item_headers, start=1):
        cell = ws_items.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    
    item_row = 4
    for o in qs:
        for it in o.items.all():
            lt = float(it.qty) * float(it.unit_price)
            ws_items.cell(row=item_row, column=1, value=o.id)
            ws_items.cell(row=item_row, column=2, value=it.id)
            ws_items.cell(row=item_row, column=3, value=it.product_id)
            ws_items.cell(row=item_row, column=4, value=it.product.name)
            ws_items.cell(row=item_row, column=5, value=it.qty)
            ws_items.cell(row=item_row, column=6, value=f"{it.unit_price:,.2f} DT")
            ws_items.cell(row=item_row, column=7, value=f"{lt:,.2f} DT")
            ws_items.cell(row=item_row, column=8, value=o.status)
            
            item_row += 1
    
    # Appliquer les styles à toutes les feuilles
    _apply_excel_styles(ws_summary)
    _apply_excel_styles(ws_products)
    _apply_excel_styles(ws_buyers)
    _apply_excel_styles(ws_regions)
    _apply_excel_styles(ws_trend)
    _apply_excel_styles(ws_orders)
    _apply_excel_styles(ws_items)
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"commandes_agricycle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@api_view(["GET"])
def export_orders_report_pdf(request):
    if not _admin_only(request.user):
        return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)

    status_filter = request.query_params.get("status")
    date_from = request.query_params.get("from")
    date_to = request.query_params.get("to")

    qs = Order.objects.all().prefetch_related("items","items__product").order_by("-created_at")
    if status_filter:
        qs = qs.filter(status=status_filter)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    revenue_qs = qs.exclude(status__in=[Order.Status.CART, Order.Status.CANCELLED])

    by_status = qs.values("status").annotate(orders_count=Count("id"), amount_sum=Sum("total_amount")).order_by("status")
    totals = qs.aggregate(orders_count=Count("id"), amount_sum=Sum("total_amount"))
    revenue_totals = revenue_qs.aggregate(revenue_orders_count=Count("id"), revenue_sum=Sum("total_amount"))

    line_total = ExpressionWrapper(F("qty") * F("unit_price"), output_field=DecimalField(max_digits=12, decimal_places=2))

    top_products = (
        OrderItem.objects.filter(order__in=revenue_qs)
        .values("product__name")
        .annotate(qty_sum=Sum("qty"), revenue_sum=Sum(line_total))
        .order_by("-revenue_sum")[:10]
    )

    top_regions = (
        revenue_qs.values("buyer__governorate")
        .annotate(orders_count=Count("id"), revenue_sum=Sum("total_amount"))
        .order_by("-revenue_sum")[:10]
    )

    # Trend monthly revenue (last 12)
    monthly = defaultdict(float)
    for o in revenue_qs:
        monthly[o.created_at.strftime("%Y-%m")] += float(o.total_amount)
    trend_rows = sorted(monthly.items())[-12:]

    resp = HttpResponse(content_type="application/pdf")
    filename = f"rapport_commandes_agricycle_{date_from or 'tout'}_{date_to or 'auj'}.pdf"
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    # Créer le canvas avec une page A4
    c = canvas.Canvas(resp, pagesize=A4)
    width, height = A4
    
    # Couleurs
    primary_color = (46/255, 139/255, 87/255)  # Vert AgriCycle
    secondary_color = (76/255, 175/255, 80/255)  # Vert plus clair
    accent_color = (255/255, 152/255, 0/255)  # Orange
    dark_gray = (55/255, 71/255, 79/255)
    light_gray = (236/255, 239/255, 241/255)
    
    def draw_header():
        """Dessiner l'en-tête avec logo et informations"""
        # Rectangle d'en-tête
        c.setFillColorRGB(*primary_color)
        c.rect(0, height - 4*cm, width, 4*cm, fill=1)
        
        # Logo et nom d'entreprise
        c.setFillColorRGB(1, 1, 1)  # Blanc
        c.setFont("Helvetica-Bold", 20)
        c.drawString(2*cm, height - 2*cm, "🛡️ AgriCycle")
        
        c.setFont("Helvetica", 10)
        c.drawString(2*cm, height - 2.8*cm, "Plateforme de Recyclage Agricole")
        
        # Informations du rapport
        c.setFont("Helvetica-Bold", 12)
        c.drawRightString(width - 2*cm, height - 2*cm, "Rapport des Commandes")
        
        c.setFont("Helvetica", 9)
        periode = f"Période: {date_from or 'Début'} au {date_to or 'Aujourd hui'}"
        c.drawRightString(width - 2*cm, height - 2.6*cm, periode)
        
        statut = f"Statut: {status_filter or 'Tous'}"
        c.drawRightString(width - 2*cm, height - 3.2*cm, statut)
        
        date_gen = f"Généré le: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        c.drawRightString(width - 2*cm, height - 3.8*cm, date_gen)
        
        return height - 4.5*cm
    
    def draw_section_title(y, title):
        """Dessiner un titre de section"""
        c.setFillColorRGB(*primary_color)
        c.setFont("Helvetica-Bold", 12)
        c.drawString(2*cm, y, title)
        
        # Ligne de séparation
        c.setStrokeColorRGB(*primary_color)
        c.setLineWidth(1)
        c.line(2*cm, y - 0.3*cm, width - 2*cm, y - 0.3*cm)
        
        return y - 0.6*cm
    
    def draw_table_header(y, headers):
        """Dessiner l'en-tête d'un tableau"""
        c.setFillColorRGB(*secondary_color)
        c.rect(2*cm, y - 0.6*cm, width - 4*cm, 0.6*cm, fill=1)
        
        c.setFillColorRGB(1, 1, 1)
        c.setFont("Helvetica-Bold", 10)
        
        # Calculer la largeur des colonnes
        col_width = (width - 4*cm) / len(headers)
        
        for i, header in enumerate(headers):
            x = 2*cm + i * col_width
            c.drawString(x + 0.2*cm, y - 0.4*cm, header)
        
        return y - 0.8*cm
    
    def draw_table_row(y, values, is_highlight=False):
        """Dessiner une ligne de tableau"""
        if is_highlight:
            c.setFillColorRGB(0.95, 0.95, 0.95)
            c.rect(2*cm, y - 0.5*cm, width - 4*cm, 0.5*cm, fill=1)
        
        c.setFillColorRGB(0, 0, 0)
        c.setFont("Helvetica", 9)
        
        col_width = (width - 4*cm) / len(values)
        
        for i, value in enumerate(values):
            x = 2*cm + i * col_width
            c.drawString(x + 0.2*cm, y - 0.3*cm, str(value))
        
        return y - 0.6*cm
    
    def new_page():
        """Créer une nouvelle page"""
        c.showPage()
        c.setFont("Helvetica", 10)
        return height - 2*cm
    
    def format_currency(amount):
        """Formater un montant en devise tunisienne"""
        return f"{float(amount or 0):,.2f} DT"
    
    # Page 1: En-tête et totaux
    current_y = draw_header()
    
    # Section 1: Totaux généraux
    current_y = draw_section_title(current_y, "📊 Totaux Généraux")
    
    # Tableau des totaux
    current_y = draw_table_header(current_y, ["Description", "Valeur"])
    
    current_y = draw_table_row(current_y, ["Toutes les commandes", totals["orders_count"] or 0])
    current_y = draw_table_row(current_y, ["Montant total", format_currency(totals["amount_sum"])], True)
    current_y = draw_table_row(current_y, ["Commandes validées (hors paniers)", revenue_totals["revenue_orders_count"] or 0])
    current_y = draw_table_row(current_y, ["Revenu généré", format_currency(revenue_totals["revenue_sum"])], True)
    
    current_y -= 0.5*cm
    
    # Section 2: Répartition par statut
    if current_y < 8*cm:
        current_y = new_page()
    
    current_y = draw_section_title(current_y, "📈 Répartition par Statut")
    
    current_y = draw_table_header(current_y, ["Statut", "Nombre de Commandes", "Montant Total (DT)"])
    
    for r in by_status:
        if current_y < 4*cm:
            current_y = new_page()
            current_y = draw_table_header(current_y, ["Statut", "Nombre de Commandes", "Montant Total (DT)"])
        
        current_y = draw_table_row(current_y, [
            r["status"],
            r["orders_count"],
            format_currency(r["amount_sum"])
        ])
    
    current_y -= 0.5*cm
    
    # Section 3: Top produits
    if current_y < 8*cm:
        current_y = new_page()
    
    current_y = draw_section_title(current_y, "🏆 Top Produits par Revenu")
    
    current_y = draw_table_header(current_y, ["Produit", "Quantité Vendue", "Revenu Généré (DT)"])
    
    for p in top_products:
        if current_y < 4*cm:
            current_y = new_page()
            current_y = draw_table_header(current_y, ["Produit", "Quantité Vendue", "Revenu Généré (DT)"])
        
        current_y = draw_table_row(current_y, [
            p["product__name"],
            p["qty_sum"] or 0,
            format_currency(p["revenue_sum"])
        ])
    
    current_y -= 0.5*cm
    
    # Section 4: Top régions
    if current_y < 8*cm:
        current_y = new_page()
    
    current_y = draw_section_title(current_y, "📍 Top Régions par Revenu")
    
    current_y = draw_table_header(current_y, ["Gouvernorat", "Nombre de Commandes", "Revenu (DT)"])
    
    for r in top_regions:
        if current_y < 4*cm:
            current_y = new_page()
            current_y = draw_table_header(current_y, ["Gouvernorat", "Nombre de Commandes", "Revenu (DT)"])
        
        current_y = draw_table_row(current_y, [
            r["buyer__governorate"] or "Non spécifié",
            r["orders_count"] or 0,
            format_currency(r["revenue_sum"])
        ])
    
    current_y -= 0.5*cm
    
    # Section 5: Tendances mensuelles
    if current_y < 8*cm:
        current_y = new_page()
    
    current_y = draw_section_title(current_y, "📅 Tendances Mensuelles (12 derniers mois)")
    
    current_y = draw_table_header(current_y, ["Mois", "Revenu (DT)"])
    
    for m, rev in trend_rows:
        if current_y < 4*cm:
            current_y = new_page()
            current_y = draw_table_header(current_y, ["Mois", "Revenu (DT)"])
        
        current_y = draw_table_row(current_y, [m, format_currency(rev)])
    
    # Pied de page
    c.setFillColorRGB(*dark_gray)
    c.setFont("Helvetica", 8)
    c.drawString(2*cm, 1.5*cm, f"© {timezone.now().year} AgriCycle - Tous droits réservés")
    c.drawRightString(width - 2*cm, 1.5*cm, f"Page 1 sur 1")
    
    # Informations de contact
    contact_y = 1*cm
    c.drawString(2*cm, contact_y, "📧 contact@agricycle.tn")
    c.drawCentredString(width/2, contact_y, "🌐 www.agricycle.tn")
    c.drawRightString(width - 2*cm, contact_y, "📞 +216 XX XXX XXX")
    
    c.save()
    return resp

# ---------- Me/Profile ----------
class MeViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsNotSuspended]

    @action(detail=False, methods=["get"])
    def me(self, request):
        return Response(UserMeSerializer(request.user).data)

    @action(detail=False, methods=["patch"])
    def update_profile(self, request):
        u = request.user
        allowed = {
            "governorate","delegation",
            "farm_name","farm_address","farm_lat","farm_lng",
        }
        for k,v in request.data.items():
            if k in allowed:
                setattr(u, k, v)
        u.save()
        return Response(UserMeSerializer(u).data)


# ---------- Declarations ----------
class DeclarationViewSet(viewsets.ModelViewSet):
    serializer_class = DeclarationSerializer
    permission_classes = [IsAuthenticated, IsNotSuspended, IsOwnerOrAdmin]
    queryset = Declaration.objects.all().select_related("owner").order_by("-created_at")

    def get_queryset(self):
        u = self.request.user
        if u.is_admin_role():
            return self.queryset
        return self.queryset.filter(owner=u)

    def perform_create(self, serializer):
        u = self.request.user
        if u.role not in {User.Role.FARMER, User.Role.DUAL} and not u.is_admin_role():
            raise PermissionError("Forbidden")
        serializer.save(owner=u, status=Declaration.Status.PENDING)

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsNotSuspended, IsAdminRole])
    def validate_declaration(self, request, pk=None):
        d = self.get_object()
        d.status = Declaration.Status.VALIDATED
        d.save(update_fields=["status"])
        AdminAuditLog.objects.create(admin=request.user, action="DECLARATION_VALIDATED", meta={"declaration_id": d.id})
        return Response(DeclarationSerializer(d).data)

    @action(detail=True, methods=["post"], permission_classes=[IsAuthenticated, IsNotSuspended, IsAdminRole])
    def mark_collected(self, request, pk=None):
        d = self.get_object()
        d.status = Declaration.Status.COLLECTED
        d.save(update_fields=["status"])
        AdminAuditLog.objects.create(admin=request.user, action="DECLARATION_COLLECTED", meta={"declaration_id": d.id})
        return Response(DeclarationSerializer(d).data)


# ---------- Products ----------
class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    queryset = Product.objects.all().order_by("-id")

    def get_permissions(self):
        if self.request.method in ["GET"]:
            # Public list allowed (frontend catalog)
            return [AllowAny()]
        return [IsAuthenticated(), IsNotSuspended(), IsAdminRole()]

    def get_queryset(self):
        u = self.request.user
        # Public only sees active
        if not u.is_authenticated or not u.is_admin_role():
            return self.queryset.filter(is_active=True)
        return self.queryset


# ---------- Orders + Cart ----------
class OrderViewSet(viewsets.ModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated, IsNotSuspended, IsOwnerOrAdmin]
    queryset = Order.objects.all().prefetch_related("items","items__product").order_by("-created_at")

    def get_queryset(self):
        u = self.request.user
        if u.is_admin_role():
            return self.queryset
        return self.queryset.filter(buyer=u)

    def perform_create(self, serializer):
        u = self.request.user
        if u.role not in {User.Role.BENEFICIARY, User.Role.DUAL} and not u.is_admin_role():
            raise PermissionError("Forbidden")
        serializer.save(buyer=u)

    @action(detail=False, methods=["get"])
    def my_cart(self, request):
        u = request.user
        if u.role not in {User.Role.BENEFICIARY, User.Role.DUAL} and not u.is_admin_role():
            return Response({"detail":"Forbidden"}, status=403)

        cart = Order.objects.filter(buyer=u, status=Order.Status.CART).order_by("-created_at").first()
        if not cart:
            cart = Order.objects.create(buyer=u, status=Order.Status.CART)
        return Response(OrderSerializer(cart).data)

    @action(detail=True, methods=["post"])
    def add_item(self, request, pk=None):
        order = self.get_object()
        if order.status != Order.Status.CART and not request.user.is_admin_role():
            return Response({"detail":"Only CART can be modified"}, status=400)

        ser = CartAddItemSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        product = get_object_or_404(Product, id=ser.validated_data["product_id"], is_active=True)

        qty = ser.validated_data["qty"]
        item, created = OrderItem.objects.get_or_create(
            order=order, product=product,
            defaults={"qty": qty, "unit_price": product.price}
        )
        if not created:
            item.qty = item.qty + qty
            item.unit_price = product.price
            item.save(update_fields=["qty","unit_price"])
        order.recalc_total()
        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=["patch"], url_path=r"items/(?P<item_id>\d+)")
    def update_item(self, request, pk=None, item_id=None):
        order = self.get_object()
        if order.status != Order.Status.CART and not request.user.is_admin_role():
            return Response({"detail":"Only CART can be modified"}, status=400)

        ser = CartUpdateItemSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        item = get_object_or_404(OrderItem, id=item_id, order=order)
        item.qty = ser.validated_data["qty"]
        item.save(update_fields=["qty"])
        order.recalc_total()
        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=["delete"], url_path=r"items/(?P<item_id>\d+)")
    def remove_item(self, request, pk=None, item_id=None):
        order = self.get_object()
        if order.status != Order.Status.CART and not request.user.is_admin_role():
            return Response({"detail":"Only CART can be modified"}, status=400)
        item = get_object_or_404(OrderItem, id=item_id, order=order)
        item.delete()
        order.recalc_total()
        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=["post"])
    def checkout(self, request, pk=None):
        order = self.get_object()
        if not request.user.is_admin_role() and order.buyer_id != request.user.id:
            return Response({"detail":"Forbidden"}, status=403)
        if order.status != Order.Status.CART:
            return Response({"detail":"Not a cart"}, status=400)
        if order.items.count() == 0:
            return Response({"detail":"Cart empty"}, status=400)
        order.status = Order.Status.CONFIRMED
        order.save(update_fields=["status"])
        return Response(OrderSerializer(order).data)

    @action(detail=True, methods=["post"])
    def set_status(self, request, pk=None):
        if not request.user.is_admin_role():
            return Response({"detail":"Forbidden"}, status=403)
        order = self.get_object()
        ser = OrderSetStatusSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        order.status = ser.validated_data["status"]
        order.save(update_fields=["status"])
        AdminAuditLog.objects.create(admin=request.user, action="ORDER_SET_STATUS", meta={"order_id": order.id, "status": order.status})
        return Response(OrderSerializer(order).data)


# ---------- Good Practices ----------
class GoodPracticeViewSet(viewsets.ModelViewSet):
    serializer_class = GoodPracticeSerializer
    queryset = GoodPractice.objects.all().order_by("-created_at")

    def get_permissions(self):
        if self.request.method in ["GET"]:
            return [AllowAny()]
        return [IsAuthenticated(), IsNotSuspended(), IsAdminRole()]


# ---------- User ViewSet ----------
class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]

    # Méthode pour supprimer un utilisateur
    @action(detail=True, methods=["delete"])
    def delete_user(self, request, pk=None):
        user = self.get_object()
        user.delete()
        return Response(status=204)